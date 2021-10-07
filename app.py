from flask import Flask, render_template, redirect, request, session, url_for, send_file
import sqlite3
from datetime import datetime, timedelta, date
from werkzeug.security import check_password_hash, generate_password_hash
from io import BytesIO
import openpyxl as xl
from openpyxl.styles import Font
from os import path

app = Flask(__name__)
app.secret_key = "hello"
app.config['SESSION_TYPE'] = 'filesystem'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=5)

root = path.dirname(path.realpath(__file__))

def drcr(amount):
    if amount >= 0:
        return f"{amount:,.2f} Dr."

    else:
        return f"{abs(amount):,.2f} Cr."

def toabs(amount):
    return f"{abs(amount):,.2f}"

def topty(amount):
    if amount == 0:
        return "0.00"
    elif amount > 0:
        return f"{amount:,.2f}"
    else:
        return f"({- amount:,.2f})"

def tomillion(amount):
    m = int(amount / 1000000)

    return f"{m:,} m"

types = {
        "NCA": "Non-current asset",
        "CA": "Current asset",
        "NCL": "Non-current liability",
        "CL": "Current liability",
        "EQT": "Equity",
        "INC": "Income",
        "EXP": "Expenses"
    }

subtypes = {
    "ppe": "Property, plant and equipment",
    "investments": "Investments",
    "intangible": "Intangibles",
    "inventories": "Inventories",
    "receivables": "Trade receivables",
    "cash": "Cash and cash equivalents",
    "long-borrowings": "Long-term borrowings",
    "deferred-tax": "Deferred tax",
    "payables": "Trade and other payables",
    "short-borrowings": "Short term borrowings",
    "tax-payable": "Current tax payable",
    "provisions": "Short-term provisions",
    "capital": "Capital",
    "other-equity": "Other components of equity",
    "sales": "Sales",
    "investment-income": "Investment income",
    "other-income": "Other income",
    "cost-of-sales": "Cost of sales",
    "distribution-costs": "Distribution costs",
    "admin-exp": "Administrative expenses",
    "finance-costs": "Finance costs",
    "tax-exp": "Income tax expense",
}

currency_list = ["RM", "$", "€", "£", "¥"]

@app.route("/")
def index():
    if "name" in session:
        return redirect("/home")

    else:
        return render_template("index.html")

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        name = request.form.get("login-name")
        password = request.form.get("login-password")

        if not name or not password:
            return render_template("login.html", msg="Input field is empty")

        conn = sqlite3.connect(path.join(root, "data.db"))
        db = conn.cursor()
        db.execute("SELECT password FROM persons WHERE name=?", (name, ))
        password_list = db.fetchall()
        
        if len(password_list) != 1:
            conn.close()
            return render_template("login.html", msg="Invalid username")
        
        if check_password_hash(password_list[0][0], password): 
            conn.close()
            session.permanent = True
            session["name"] = name
            return redirect("/home")

        else:
            conn.close()
            return render_template("login.html", msg="Invalid password")

    else:
        return render_template("login.html")

@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "POST":
        name = request.form.get("signup-name")
        password1 = request.form.get("signup-password1")
        password2 = request.form.get("signup-password2")
        currency = request.form.get("currency")
        print(currency)
        date = datetime.now().replace(microsecond=0)
        agree = request.form.get("agree")

        if not name or not password1 or not password2:
            return render_template("signup.html", msg="Input field is empty", currency_list=currency_list)

        if password1 != password2:
            return render_template("signup.html", msg="Password and confirmation password are different", currency_list=currency_list)

        if agree != "agree":
            return render_template("signup.html", msg="You must agree the terms of use to sign up", currency_list=currency_list)

        if currency not in currency_list:
            return render_template("signup.html", msg="Invalid currency", currency_list=currency_list)

        conn = sqlite3.connect(path.join(root, "data.db"))
        db = conn.cursor()
        db.execute("SELECT * FROM persons WHERE name = ?", (name, ))
        if len(db.fetchall()) == 1:
            conn.close()
            return render_template("signup.html", msg="Username has been taken. Please choose another username.", currency_list=currency_list)
        else:
            db.execute("INSERT INTO persons (name, password, date, currency) VALUES (?, ?, ?, ?)", (name, generate_password_hash(password1), date, currency))
            db.execute("SELECT id FROM persons WHERE name = ?", (name, ))
            id = db.fetchall()[0][0]
            db.execute("INSERT INTO accounts (name, type, subtype, balance, persons_id, dependency, deleted) VALUES (?, ?, ?, ?, ?, ?, ?)", ("Cash", "CA", "cash", 0, id, 0, 0))
            db.execute("INSERT INTO accounts (name, type, subtype, balance, persons_id, dependency, deleted) VALUES (?, ?, ?, ?, ?, ?, ?)", ("Bank", "CA", "cash", 0, id, 0, 0))
            db.execute("INSERT INTO accounts (name, type, subtype, balance, persons_id, dependency, deleted) VALUES (?, ?, ?, ?, ?, ?, ?)", ("Capital", "EQT", "capital", 0, id, 0, 0))
            conn.commit()
            conn.close()

            session.permanent = True
            session["name"] = name
            return redirect("/home")

    else:
        return render_template("signup.html", currency_list=currency_list)


@app.route("/home", methods=["GET", "POST"])
def home():
    if "name" not in session:
        return redirect("/")

    if request.method == "POST":
        person = session["name"]
        debit = request.form.get("debit")
        credit = request.form.get("credit")
        particular = request.form.get("particular")
        amount = request.form.get("amount")
        date = datetime.now().replace(microsecond=0)

        if not debit or not credit or not particular or not amount:
            return redirect("/home")

        if debit == credit:
            return redirect("/home")

        conn = sqlite3.connect(path.join(root, "data.db"))
        db = conn.cursor()          

        db.execute("SELECT id FROM persons WHERE name = ?", (person, ))
        id = db.fetchall()[0][0]

        db.execute("SELECT id FROM accounts WHERE persons_id = ? AND name = ?", (id, debit))
        debit_id = db.fetchall()[0][0]

        db.execute("SELECT id FROM accounts WHERE persons_id = ? AND name = ?", (id, credit))
        credit_id = db.fetchall()[0][0]

        db.execute("INSERT INTO transactions (persons_id, debit_id, credit_id, particular, amount, date) VALUES (?, ?, ?, ?, ?, ?)", (id, debit_id, credit_id, particular, amount, date))
        db.execute("SELECT balance FROM accounts WHERE id = ?", (debit_id, ))
        debit_balance = db.fetchall()[0][0]
        db.execute("SELECT balance FROM accounts WHERE id = ?", (credit_id, ))
        credit_balance = db.fetchall()[0][0]
        db.execute("UPDATE accounts SET balance = ? WHERE persons_id = ? AND id = ?", (debit_balance + float(amount), id, debit_id))
        db.execute("UPDATE accounts SET balance = ? WHERE persons_id = ? AND id = ?", (credit_balance - float(amount), id, credit_id))

        conn.commit()
        conn.close()
        return redirect("/home")

    else:
        person = session["name"]
        conn = sqlite3.connect(path.join(root, "data.db"))
        db = conn.cursor()
        db.execute("SELECT name FROM accounts WHERE persons_id = (SELECT id FROM persons WHERE name = ?) AND deleted = 0", (person, ))
        accounts = [item[0] for item in db.fetchall()]
        db.execute('SELECT SUM(balance) FROM accounts WHERE subtype = "cash" AND persons_id = (SELECT id FROM persons WHERE name = ?) AND deleted = 0', (person, ))
        balance = db.fetchall()[0][0]

        db.execute("SELECT currency FROM persons WHERE name = ?", (person, ))
        currency = db.fetchall()[0][0]

        if not balance:
            balance = 0

        if balance > 1000000 or balance < - 1000000:
            balance = tomillion(balance)
        else:
            balance = topty(balance)

        db.execute('SELECT SUM(balance) FROM accounts WHERE type = "INC" AND persons_id = (SELECT id FROM persons WHERE name = ?) AND deleted = 0', (person, ))
        income = db.fetchall()[0][0]

        db.execute('SELECT SUM(balance) FROM accounts WHERE type = "EXP" AND persons_id = (SELECT id FROM persons WHERE name = ?) AND deleted = 0', (person, ))
        expenses = db.fetchall()[0][0]

        if not income:
            income = 0

        if not expenses:
            expenses = 0

        pft = - (income + expenses)

        if pft > 1000000 or pft < - 1000000:
            profit = tomillion(pft)
        else:
            profit = topty(pft)

        transactions = []

        db.execute("SELECT * FROM transactions WHERE persons_id = (SELECT id FROM persons WHERE name = ?)", (person, ))
        history = db.fetchall()

        for item in history:
            date = item[6]

            db.execute("SELECT name FROM accounts WHERE id = ?", (item[2], ))
            debit = db.fetchall()[0][0]

            db.execute("SELECT name FROM accounts WHERE id = ?", (item[3], ))
            credit = db.fetchall()[0][0]

            particular = item[4]
            amount = toabs(item[5])

            transactions.append([date, debit, credit, particular, amount])

        conn.close()
        return render_template("home.html", accounts=accounts, balance=balance, profit=profit, pft=pft, transactions=transactions, currency=currency)

@app.route("/add-account", methods=["GET", "POST"])
def add_account():
    if "name" not in session:
        return redirect("/")

    if request.method == "POST":
        person = session["name"]
        type = request.form.get("type")
        subtype = request.form.get("subtype")
        account_name = request.form.get("account-name")

        if type not in types or not subtype or not account_name:
            return render_template("add-account.html", types=list(types.items()), alert_msg="Input field is empty")

        conn = sqlite3.connect(path.join(root, "data.db"))
        db = conn.cursor()
        db.execute("SELECT id FROM persons WHERE name = ?", (person, ))

        id = db.fetchall()[0][0]
        db.execute("SELECT id, deleted FROM accounts WHERE name = ? AND persons_id = ?", (account_name, id))
        record = db.fetchall()

        if len(record) != 1:
            db.execute("INSERT INTO accounts (name, type, subtype, balance, persons_id, dependency, deleted) VALUES (?, ?, ?, ?, ?, ?, ?)", (account_name, type, subtype, 0, id, 0, 0))
            conn.commit()
            conn.close()
            return render_template("add-account.html", types=list(types.items()), primary_msg="Account added")

        if len(record) == 1 and record[0][1] == 1:
            db.execute("UPDATE accounts SET deleted = 0 WHERE id = ?", (record[0][0], ))
            conn.commit()
            conn.close()
            return render_template("add-account.html", types=list(types.items()), primary_msg="Archived account recovered")


        else:
            conn.close()
            return render_template("add-account.html", types=list(types.items()), alert_msg="Account name must be unique")

    else:
        return render_template("add-account.html", types=list(types.items()))

@app.route("/view-account")
def view_account():
    if "name" not in session:
        return redirect("/")

    success = request.args.get("success", None)

    person = session["name"]

    conn = sqlite3.connect(path.join(root, "data.db"))
    db = conn.cursor()
    db.execute("SELECT * FROM accounts WHERE persons_id = (SELECT id FROM persons WHERE name = ?) AND deleted = 0", (person, ))
    accounts = db.fetchall()

    db.execute("SELECT currency FROM persons WHERE name = ?", (person, ))
    currency = db.fetchall()[0][0]

    balances = []

    for account in accounts:
        type = types[account[2]]
        subtype = subtypes[account[3]]
        balances.append([account[0], account[1], type, subtype, drcr(account[4])])
        
    conn.close()

    return render_template("view-account.html", balances=balances, success=success, currency=currency)

@app.route("/accounts/<int:id>")
def details(id):
    if "name" not in session:
        return redirect("/")

    name = session["name"]

    conn = sqlite3.connect(path.join(root, "data.db"))
    db = conn.cursor()
    db.execute("SELECT * FROM accounts WHERE id = ?", (id, ))
    account = db.fetchall()[0]

    db.execute("SELECT currency FROM persons WHERE name = ?", (name, ))
    currency = db.fetchall()[0][0]

    persons_id1 = account[5]

    db.execute("SELECT id FROM persons WHERE name = ?", (name, ))
    persons_id2 = db.fetchall()[0][0]

    if persons_id1 != persons_id2:
        conn.close()
        return redirect("/")

    account_details = [account[1], types[account[2]], subtypes[account[3]], drcr(account[4])]

    transactions = []

    db.execute("SELECT * FROM transactions WHERE debit_id = ? OR credit_id = ?", (id, id))
    history = db.fetchall()

    for item in history:
        date = item[6]

        db.execute("SELECT name FROM accounts WHERE id = ?", (item[2], ))
        debit = db.fetchall()[0][0]

        db.execute("SELECT name FROM accounts WHERE id = ?", (item[3], ))
        credit = db.fetchall()[0][0]

        particular = item[4]
        amount = topty(item[5])

        transactions.append([date, debit, credit, particular, amount])
    
    conn.commit()
    conn.close()

    return render_template("details.html", id=id, account_details=account_details, transactions=transactions, currency=currency)

@app.route("/delete-account/<int:id>")
def delete_account(id):
    if "name" not in session:
        return redirect("/")

    name = session["name"]

    conn = sqlite3.connect(path.join(root, "data.db"))
    db = conn.cursor()
    db.execute("SELECT persons_id, balance FROM accounts WHERE id = ?", (id, ))
    persons_id1, balance = db.fetchall()[0]

    db.execute("SELECT id FROM persons WHERE name = ?", (name, ))
    persons_id2 = db.fetchall()[0][0]

    if persons_id1 != persons_id2:
        conn.close()
        return redirect("/")

    if balance != 0:
        conn.close()
        return redirect(url_for("view_account", success="false"))

    else:
        db.execute("UPDATE accounts SET deleted = 1 WHERE id = ?", (id, ))
        conn.commit()
        conn.close()
        return redirect(url_for("view_account", success="true"))

@app.route("/terms-of-use")
def terms():
    return render_template("terms.html")

@app.route("/trial-balance")
def tb():
    if "name" not in session:
        return redirect("/")

    person = session["name"]

    conn = sqlite3.connect(path.join(root, "data.db"))
    db = conn.cursor()

    db.execute("SELECT name, balance FROM accounts WHERE persons_id = (SELECT id FROM persons WHERE name = ?) AND deleted = 0", (person, ))
    balances = db.fetchall()

    db.execute("SELECT currency FROM persons WHERE name = ?", (person, ))
    currency = db.fetchall()[0][0]

    today_date = date.today().strftime('%d %B %Y')

    accounts = []

    for balance in balances:
        if balance[1] >= 0:
            accounts.append([balance[0], toabs(balance[1]), None])
        else:
            accounts.append([balance[0], None, toabs(balance[1])])

    db.execute("SELECT SUM(balance) FROM accounts WHERE persons_id = (SELECT id FROM persons WHERE name = ?) AND deleted = 0 AND balance >= 0", (person, ))
    debit_total = toabs(db.fetchall()[0][0])

    db.execute("SELECT SUM(balance) FROM accounts WHERE persons_id = (SELECT id FROM persons WHERE name = ?) AND deleted = 0 AND balance >= 0", (person, ))
    credit_total = toabs(db.fetchall()[0][0])
            
    conn.close()
    return render_template("trial-balance.html", accounts=accounts, debit_total=debit_total, credit_total=credit_total, today_date=today_date, currency=currency)

@app.route("/sopl")
def sopl():
    if "name" not in session:
        return redirect("/")

    year = request.args.get("year", "all")

    person = session["name"]

    conn = sqlite3.connect(path.join(root, "data.db"))
    db = conn.cursor()

    db.execute("SELECT currency FROM persons WHERE name = ?", (person, ))
    currency = db.fetchall()[0][0]

    db.execute("SELECT id FROM persons WHERE name = ?", (person, ))
    id = db.fetchall()[0][0]

    db.execute("SELECT DISTINCT(strftime('%Y', date)) from transactions WHERE persons_id = ?", (id, ))
    yrs = db.fetchall()

    if year not in [str(yr[0]) for yr in yrs] and year != "all":
        return redirect("/")

    sopl_list = []

    for subtype in list(subtypes.keys())[14:]:
        if year == "all":
            db.execute("SELECT SUM(amount) FROM transactions WHERE debit_id IN (SELECT id FROM accounts WHERE subtype = ?) AND persons_id = ?", (subtype, id))
            debit = db.fetchall()[0][0]

            db.execute("SELECT SUM(amount) FROM transactions WHERE credit_id IN (SELECT id FROM accounts WHERE subtype = ?) AND persons_id = ?", (subtype, id))
            credit = db.fetchall()[0][0]

        else:
            db.execute("SELECT SUM(amount) FROM transactions WHERE debit_id IN (SELECT id FROM accounts WHERE subtype = ?) AND persons_id = ? AND strftime('%Y', date) = ?", (subtype, id, year))
            debit = db.fetchall()[0][0]

            db.execute("SELECT SUM(amount) FROM transactions WHERE credit_id IN (SELECT id FROM accounts WHERE subtype = ?) AND persons_id = ? AND strftime('%Y', date) = ?", (subtype, id, year))
            credit = db.fetchall()[0][0]

        if debit is None:
            debit = 0

        if credit is None:
            credit = 0

        amount = credit - debit
        sopl_list.append([subtype, amount])

    gp = sopl_list[0][1] + sopl_list[3][1]
    pfo = gp + sopl_list[4][1] + sopl_list[5][1] + sopl_list[2][1]
    pbt = pfo + sopl_list[6][1] + sopl_list[1][1]
    pat = pbt + sopl_list[7][1]

    total = [topty(gp), topty(pfo), topty(pbt), topty(pat)]

    for i in range(len(sopl_list)):
        sopl_list[i][1] = topty(sopl_list[i][1])

    today_date = date.today().strftime('%d %B %Y')
    today_year = date.today().strftime('%Y')

    conn.close()
    return render_template("sopl.html", today_date=today_date, sopl_list=sopl_list, total=total, subtypes=subtypes, year=year, yrs=yrs, today_year=today_year, currency=currency)

@app.route("/sofp")
def sofp():
    if "name" not in session:
        return redirect("/")

    person = session["name"]

    today_date = date.today().strftime('%d %B %Y')

    conn = sqlite3.connect(path.join(root, "data.db"))
    db = conn.cursor()

    db.execute("SELECT currency FROM persons WHERE name = ?", (person, ))
    currency = db.fetchall()[0][0]

    sofp_list = []

    for subtype in list(subtypes.keys())[:14]:
        db.execute("SELECT SUM(balance) FROM accounts WHERE persons_id = (SELECT id FROM persons WHERE name = ?) AND deleted = 0 AND subtype = ?", (person, subtype))
        balance = db.fetchall()[0][0]

        if balance is None:
            balance = 0.0

        sofp_list.append([subtype, balance])

    db.execute('SELECT SUM(balance) FROM accounts WHERE type = "INC" AND persons_id = (SELECT id FROM persons WHERE name = ?) AND deleted = 0', (person, ))
    income = db.fetchall()[0][0]

    db.execute('SELECT SUM(balance) FROM accounts WHERE type = "EXP" AND persons_id = (SELECT id FROM persons WHERE name = ?) AND deleted = 0', (person, ))
    expenses = db.fetchall()[0][0]

    if not income:
        income = 0

    if not expenses:
        expenses = 0

    profit = income + expenses

    sofp_list[12][1] = sofp_list[12][1] + profit

    # Handle overdraft

    if sofp_list[5][1] < 0:
        sofp_list[8][1] = sofp_list[8][1] - sofp_list[12][1]
        sofp_list[5][1] = 0.0

    nca = sofp_list[0][1] + sofp_list[1][1] + sofp_list[2][1]
    ca = sofp_list[3][1] + sofp_list[4][1] + sofp_list[5][1]
    ncl = sofp_list[6][1] + sofp_list[7][1]
    cl = sofp_list[8][1] + sofp_list[9][1] + sofp_list[10][1] + sofp_list[11][1]
    eqt = sofp_list[12][1] + sofp_list[13][1]

    ast = nca + ca
    liaeqt = ncl + cl + eqt

    total = [topty(nca), topty(ca), topty(ast), topty(- eqt), topty(- ncl), topty(- cl), topty(- liaeqt)]

    for i in range(6):
        sofp_list[i][1] = topty(sofp_list[i][1])

    for i in range(6, 14):
        sofp_list[i][1] = topty(- sofp_list[i][1])

    conn.close()
    return render_template("sofp.html", sofp_list=sofp_list, subtypes=subtypes, total=total, today_date=today_date, currency=currency)

@app.route("/trial-balance/download-excel")
def tb_download_excel():
    if "name" not in session:
        return redirect("/")

    person = session["name"]

    conn = sqlite3.connect(path.join(root, "data.db"))
    db = conn.cursor()

    db.execute("SELECT currency FROM persons WHERE name = ?", (person, ))
    currency = db.fetchall()[0][0]

    db.execute("SELECT name, balance FROM accounts WHERE persons_id = (SELECT id FROM persons WHERE name = ?) AND deleted = 0", (person, ))
    balances = db.fetchall()

    output = BytesIO()
    wb = xl.Workbook()
    sheet = wb.active
    sheet.title = "Trial balance"

    boldfont = Font(bold=True)

    today_date = date.today().strftime('%d %B %Y')

    sheet.cell(1, 1).value = person
    sheet.cell(2, 1).value = f"Trial Balance as at {today_date}"

    sheet.cell(3, 1).value = "Account name"
    sheet.cell(3, 2).value = f"Debit ({currency})"
    sheet.cell(3, 3).value = f"Credit ({currency})"

    sheet.cell(3, 1).font = boldfont
    sheet.cell(3, 2).font = boldfont    
    sheet.cell(3, 3).font = boldfont

    for index, balance in enumerate(balances):
        name_cell = sheet.cell(index + 4, 1)
        name_cell.value = balance[0]

        if balance[1] >= 0:
            sheet.cell(index + 4, 2).value = abs(balance[1])
        else:
            sheet.cell(index + 4, 3).value = abs(balance[1])

    sheet.cell(len(balances) + 4, 2).value = f"=SUM(B3:B{len(balances) + 3})"
    sheet.cell(len(balances) + 4, 3).value = f"=SUM(C3:C{len(balances) + 3})"

    sheet.cell(len(balances) + 4, 2).font = boldfont
    sheet.cell(len(balances) + 4, 3).font = boldfont

    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10

    wb.save(output)
    output.seek(0)

    conn.close()
    return send_file(output, download_name="trial-balance.xlsx", as_attachment=True)

@app.route("/sopl/download-excel")
def sopl_download_excel():
    if "name" not in session:
        return redirect("/")

    year = request.args.get("year", "all")

    person = session["name"]

    conn = sqlite3.connect(path.join(root, "data.db"))
    db = conn.cursor()

    db.execute("SELECT id FROM persons WHERE name = ?", (person, ))
    id = db.fetchall()[0][0]

    db.execute("SELECT currency FROM persons WHERE name = ?", (person, ))
    currency = db.fetchall()[0][0]

    db.execute("SELECT DISTINCT(strftime('%Y', date)) from transactions WHERE persons_id = ?", (id, ))
    yrs = db.fetchall()

    if year not in [str(yr[0]) for yr in yrs] and year != "all":
        return redirect("/")

    sopl_list = []

    for subtype in list(subtypes.keys())[14:]:
        if year == "all":
            db.execute("SELECT SUM(amount) FROM transactions WHERE debit_id IN (SELECT id FROM accounts WHERE subtype = ?) AND persons_id = ?", (subtype, id))
            debit = db.fetchall()[0][0]

            db.execute("SELECT SUM(amount) FROM transactions WHERE credit_id IN (SELECT id FROM accounts WHERE subtype = ?) AND persons_id = ?", (subtype, id))
            credit = db.fetchall()[0][0]

        else:
            db.execute("SELECT SUM(amount) FROM transactions WHERE debit_id IN (SELECT id FROM accounts WHERE subtype = ?) AND persons_id = ? AND strftime('%Y', date) = ?", (subtype, id, year))
            debit = db.fetchall()[0][0]

            db.execute("SELECT SUM(amount) FROM transactions WHERE credit_id IN (SELECT id FROM accounts WHERE subtype = ?) AND persons_id = ? AND strftime('%Y', date) = ?", (subtype, id, year))
            credit = db.fetchall()[0][0]

        if debit is None:
            debit = 0

        if credit is None:
            credit = 0

        amount = credit - debit
        sopl_list.append([subtype, amount])

    today_date = date.today().strftime('%d %B %Y')
    today_year = date.today().strftime('%Y')

    output = BytesIO()
    wb = xl.Workbook()
    sheet = wb.active
    sheet.title = "SOPL"

    boldfont = Font(bold=True)
    
    sheet.cell(1, 1).value = person
    
    if year == "all" or year == today_year:
        sheet.cell(2, 1).value = f"Statement of Profit or Loss for the year ended {today_date}"
    else:
        sheet.cell(2, 1).value = f"Statement of Profit or Loss for the year ended 31 December {year}"

    sheet.cell(3, 2).value = f"{currency}"
    sheet.cell(3, 2).font = boldfont

    sheet.cell(4, 1).value = subtypes[sopl_list[0][0]]
    sheet.cell(4, 2).value = sopl_list[0][1]

    sheet.cell(5, 1).value = subtypes[sopl_list[3][0]]
    sheet.cell(5, 2).value = sopl_list[3][1]

    sheet.cell(6, 1).value = "Gross profit"
    sheet.cell(6, 2).value = "=SUM(B4:B5)"

    sheet.cell(6, 1).font = boldfont
    sheet.cell(6, 2).font = boldfont

    sheet.cell(7, 1).value = subtypes[sopl_list[4][0]]
    sheet.cell(7, 2).value = sopl_list[4][1]

    sheet.cell(8, 1).value = subtypes[sopl_list[5][0]]
    sheet.cell(8, 2).value = sopl_list[5][1]

    sheet.cell(9, 1).value = subtypes[sopl_list[2][0]]
    sheet.cell(9, 2).value = sopl_list[2][1]

    sheet.cell(10, 1).value = "Profit from operations"
    sheet.cell(10, 2).value = "=SUM(B6:B9)"

    sheet.cell(10, 1).font = boldfont
    sheet.cell(10, 2).font = boldfont

    sheet.cell(11, 1).value = subtypes[sopl_list[6][0]]
    sheet.cell(11, 2).value = sopl_list[6][1]

    sheet.cell(12, 1).value = subtypes[sopl_list[1][0]]
    sheet.cell(12, 2).value = sopl_list[1][1]

    sheet.cell(13, 1).value = "Profit before tax"
    sheet.cell(13, 2).value = "=SUM(B10:B12)"

    sheet.cell(13, 1).font = boldfont
    sheet.cell(13, 2).font = boldfont

    sheet.cell(14, 1).value = subtypes[sopl_list[7][0]]
    sheet.cell(14, 2).value = sopl_list[7][1]

    sheet.cell(15, 1).value = "Profit before tax"
    sheet.cell(15, 2).value = "=SUM(B13:B14)"

    sheet.cell(15, 1).font = boldfont
    sheet.cell(15, 2).font = boldfont

    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10

    wb.save(output)
    output.seek(0)

    conn.close()
    return send_file(output, download_name="sopl.xlsx", as_attachment=True)

@app.route("/sofp/download-excel")
def sofp_download_excel():
    if "name" not in session:
        return redirect("/")

    person = session["name"]

    today_date = date.today().strftime('%d %B %Y')

    conn = sqlite3.connect(path.join(root, "data.db"))
    db = conn.cursor()

    db.execute("SELECT currency FROM persons WHERE name = ?", (person, ))
    currency = db.fetchall()[0][0]

    sofp_list = []

    for subtype in list(subtypes.keys())[:14]:
        db.execute("SELECT SUM(balance) FROM accounts WHERE persons_id = (SELECT id FROM persons WHERE name = ?) AND deleted = 0 AND subtype = ?", (person, subtype))
        balance = db.fetchall()[0][0]

        if balance is None:
            balance = 0.0

        sofp_list.append([subtype, balance])

    db.execute('SELECT SUM(balance) FROM accounts WHERE type = "INC" AND persons_id = (SELECT id FROM persons WHERE name = ?) AND deleted = 0', (person, ))
    income = db.fetchall()[0][0]

    db.execute('SELECT SUM(balance) FROM accounts WHERE type = "EXP" AND persons_id = (SELECT id FROM persons WHERE name = ?) AND deleted = 0', (person, ))
    expenses = db.fetchall()[0][0]

    if not income:
        income = 0

    if not expenses:
        expenses = 0

    profit = income + expenses

    sofp_list[12][1] = sofp_list[12][1] + profit

    # Handle overdraft

    if sofp_list[5][1] < 0:
        sofp_list[8][1] = sofp_list[8][1] - sofp_list[12][1]
        sofp_list[5][1] = 0.0

    output = BytesIO()
    wb = xl.Workbook()
    sheet = wb.active
    sheet.title = "SOFP"

    boldfont = Font(bold=True)

    today_date = date.today().strftime('%d %B %Y')

    sheet.cell(1, 1).value = person
    sheet.cell(2, 1).value = f"Statement of Financial Position as at {today_date}"

    sheet.cell(3, 2).value = f"{currency}"
    sheet.cell(3, 3).value = f"{currency}"

    sheet.cell(3, 2).font = boldfont    
    sheet.cell(3, 3).font = boldfont

    sheet.cell(4, 1).value = "Assets"
    sheet.cell(4, 1).font = boldfont

    sheet.cell(5, 1).value = "Non-current assets"
    sheet.cell(5, 1).font = boldfont

    sheet.cell(6, 1).value = subtypes[sofp_list[0][0]]
    sheet.cell(6, 2).value = sofp_list[0][1]

    sheet.cell(7, 1).value = subtypes[sofp_list[1][0]]
    sheet.cell(7, 2).value = sofp_list[1][1]

    sheet.cell(8, 1).value = subtypes[sofp_list[2][0]]
    sheet.cell(8, 2).value = sofp_list[2][1]

    sheet.cell(9, 3).value = "=SUM(B6:B8)"

    sheet.cell(11, 1).value = "Current assets"
    sheet.cell(11, 1).font = boldfont

    sheet.cell(12, 1).value = subtypes[sofp_list[3][0]]
    sheet.cell(12, 2).value = sofp_list[3][1]

    sheet.cell(13, 1).value = subtypes[sofp_list[4][0]]
    sheet.cell(13, 2).value = sofp_list[4][1]

    sheet.cell(14, 1).value = subtypes[sofp_list[5][0]]
    sheet.cell(14, 2).value = sofp_list[5][1]

    sheet.cell(15, 3).value = "=SUM(B12:B14)"

    sheet.cell(16, 1).value = "Total assets"
    sheet.cell(16, 3).value = "=SUM(C9:C15)"

    sheet.cell(16, 1).font = boldfont
    sheet.cell(16, 3).font = boldfont

    sheet.cell(18, 1).value = "Equity and liabilities"
    sheet.cell(18, 1).font = boldfont

    sheet.cell(19, 1).value = "Capital:"
    sheet.cell(19, 1).font = boldfont

    sheet.cell(20, 1).value = subtypes[sofp_list[12][0]]
    sheet.cell(20, 2).value = - sofp_list[12][1]

    sheet.cell(21, 1).value = subtypes[sofp_list[13][0]]
    sheet.cell(21, 2).value = - sofp_list[13][1]

    sheet.cell(22, 3).value = "=SUM(B20:B21)"

    sheet.cell(24, 1).value = "Non-current liabilities"
    sheet.cell(24, 1).font = boldfont

    sheet.cell(25, 1).value = subtypes[sofp_list[6][0]]
    sheet.cell(25, 2).value = - sofp_list[6][1]

    sheet.cell(26, 1).value = subtypes[sofp_list[7][0]]
    sheet.cell(26, 2).value = - sofp_list[7][1]

    sheet.cell(27, 3).value = "=SUM(B25:B26)"

    sheet.cell(29, 1).value = "Current liabilities"
    sheet.cell(29, 1).font = boldfont

    sheet.cell(30, 1).value = subtypes[sofp_list[8][0]]
    sheet.cell(30, 2).value = - sofp_list[8][1]

    sheet.cell(31, 1).value = subtypes[sofp_list[9][0]]
    sheet.cell(31, 2).value = - sofp_list[9][1]

    sheet.cell(32, 1).value = subtypes[sofp_list[10][0]]
    sheet.cell(32, 2).value = - sofp_list[10][1]

    sheet.cell(33, 1).value = subtypes[sofp_list[11][0]]
    sheet.cell(33, 2).value = - sofp_list[11][1]

    sheet.cell(34, 3).value = "=SUM(B30:B33)"

    sheet.cell(35, 1).value = "Total equity and liabilities"
    sheet.cell(35, 3).value = "=SUM(C22:C34)"

    sheet.cell(35, 1).font = boldfont
    sheet.cell(35, 3).font = boldfont

    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10

    wb.save(output)
    output.seek(0)

    conn.close()
    return send_file(output, download_name="sofp.xlsx", as_attachment=True)

@app.route("/profile")
def profile():
    if "name" not in session:
        return redirect("/")

    conn = sqlite3.connect(path.join(root, "data.db"))
    db = conn.cursor()
    db.execute("SELECT date FROM persons WHERE name = ?", (session["name"], ))
    date = datetime.strptime(db.fetchall()[0][0], "%Y-%m-%d %H:%M:%S").strftime("%d %B %Y")

    db.execute("SELECT currency FROM persons WHERE name = ?", (session["name"], ))
    currency = db.fetchall()[0][0]

    conn.close()
    return render_template("profile.html", date=date, currency=currency)

@app.route("/change-password", methods=["POST", "GET"])
def change_password():
    if "name" not in session:
        return redirect("/")

    person = session["name"]

    if request.method == "POST":
        old_password = request.form.get("old-password")
        new_password = request.form.get("new-password")
        confirm_new_password = request.form.get("confirm-new-password")

        if not old_password or not new_password or not confirm_new_password:
            return render_template("change-password.html", alert_msg="Input field is empty")

        if new_password != confirm_new_password:
            return render_template("change-password.html", alert_msg="Password not matched")

        conn = sqlite3.connect(path.join(root, "data.db"))
        db = conn.cursor()
        db.execute("SELECT password FROM persons WHERE name=?", (person, ))
        password = db.fetchall()[0][0]
        
        if not check_password_hash(password, old_password):
            conn.close()
            return render_template("change-password.html", alert_msg="Old password is wrong")

        db.execute("UPDATE persons SET password = ? WHERE name = ?", (generate_password_hash(new_password), person))
        conn.commit()
        conn.close()

        return render_template("change-password.html", primary_msg="Password changed successfully")

    else:
        return render_template("change-password.html")

@app.route("/logout")
def logout():
    session.pop("name", None)
    return redirect("/")

if __name__ == "__main__":
    app.run()
